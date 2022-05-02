
import json
import time
import concepts
from scipy.spatial import distance
import numpy as np
import pandas as pd
import os 
from prettytable import PrettyTable
from openpyxl import Workbook
import openpyxl
from itertools import combinations
import itertools 
from functools import lru_cache
import requests
import traceback
from web3 import Web3 
from pyevolve import G1DList
from pyevolve import GSimpleGA
from pyevolve import Crossovers
from pyevolve import Mutators
from pyevolve import Selectors
from pyevolve import Initializators
from pyevolve import G2DBinaryString
from pyevolve import Consts
from pyevolve import Scaling
import copy

import argparse 

NORMAL_USER = ("NORMAL_USER")

@lru_cache(maxsize=None)
def getMethodName(hex_signature):
    # try:
    #     headers = {'X-API-KEY': 'BQYEtWs7QzCdCwkJKVhmTDp3RFTWAUEP'}
    #     url = f'https://www.4byte.directory/api/v1/signatures/?hex_signature={hex_signature}'
    #     print(url)
    #     response = requests.get(url, headers=headers)
    #     body = response.content
    #     methodName = json.loads(body.decode("utf8"))["results"][0]["text_signature"].split("(")[0]
    #     return methodName
    #     pass 
    # except:
        return hex_signature
def getABIfunctions(abi_file):
    ABIfunctions = list()
    abis = json.load(open(abi_file))
    for abi in abis:
        if abi["type"] == "function" and abi["stateMutability"]!="view":
            ABIfunctions.append(abi["name"])
    return ABIfunctions 

def getABI_file(directory):
    items = os.listdir(directory)
    abi_file = None 
    for item in items:
        if item.endswith("abi"):
            abi_file = os.path.join(directory, item)
            break 
    if abi_file:
        return abi_file
    raise Exception("ABI file not found")        

def getABIfunction_signature_mapping(abi_file):
    mappings = dict()
    abi = json.load(open(abi_file))
    signatures = [ (func['name'], '{}({})'.format(func['name'],','.join([input['type'] for input in func['inputs']]))) for func in [obj for obj in abi if obj['type'] == 'function']]
    names = set()
    for sig in signatures:
        name, full_name = sig 
        bytes4_sig = Web3.sha3(text=full_name)[0:4].hex()
        mappings[bytes4_sig] = sig
        names.add(name)

    def func(bytes4_sig):
        if bytes4_sig in names:
            return bytes4_sig
        if not bytes4_sig.startswith("0x"):
            bytes4_sig = "0x"+bytes4_sig
        if bytes4_sig in mappings:
            return mappings[bytes4_sig][0]
        else:
            return getMethodName(bytes4_sig)
    return func 
    

def lightweightrolemining(address, gene_encoding, simratio, workdir):
    print("workdir", workdir)
    starttime = time.time()
    smartcontractCalls = list()
    def getIdCounter():
            hashMap = dict()
            userMap = dict()
            def _getId(user):
                user = user.lower()
                if user in hashMap:
                    return hashMap[user]
                else:
                    hashMap[user] = len(userMap)
                    userMap[len(userMap)] = user
                    return hashMap[user]
            return userMap, _getId 
    try:
        print(os.path.join(workdir, address))
        print(getABI_file(os.path.join(workdir, address)))
        bytes4_mapping_func = getABIfunction_signature_mapping(getABI_file(os.path.join(workdir, address)))
        print("loaded abi.")
    except:
        traceback.print_exc()
        bytes4_mapping_func = getMethodName
        pass     
    with open(os.path.join(workdir, address, "user_all2.json"), "r") as f:
        usercall_statistics = json.load(f)
    UF = list()
    functions = set()
    userMap, IdCounter = getIdCounter()
    for usercall in usercall_statistics["data"]["ethereum"]["smartContractCalls"]:
        caller, count, function, success = usercall["caller"]["address"], usercall["count"], usercall["smartContractMethod"]["name"] if usercall["smartContractMethod"]["name"]!=None else \
             bytes4_mapping_func(usercall["smartContractMethod"]["signatureHash"]),  usercall["success"]
        smartcontractCalls.append((caller, function, count, success))
        if function !="Contract Creation" and (success == 1 or success is True):
            functions.add(function)
            UF.append((IdCounter(caller), function, count))
    
    functions = list(functions)
    print(len(functions), " functions", functions)
    userLabels = [(user_id, userMap[user_id]) for user_id in userMap.keys()]
    print(len(userLabels), " users")
    freqMatrix = np.zeros((len(userLabels), len(functions)))
    freqMatrix = pd.DataFrame(data=freqMatrix, columns=functions,  index=range(len(userLabels)))  
    for user_func_count in UF:
        freqMatrix.loc[user_func_count[0], user_func_count[1]] = \
            freqMatrix.loc[user_func_count[0], user_func_count[1]] + user_func_count[2]
    
    permissionMatrix = freqMatrix.astype(bool).astype(int)
    print(f"Timecost for loading history: {time.time()-starttime}")

    roleminer = GA_RM(permissionMatrix, freqMatrix, simratio, userLabels, \
        userMap, gene_encoding=gene_encoding, address=address)      
    return functions, roleminer.verybasic_usergroups, roleminer.process()


hierarchyMatrix = None 
def buildRoleHierarchy(roles):
    global hierarchyMatrix
    hierarchyMatrix = np.zeros([len(roles), len(roles)])
    for i in range(len(roles)):
        for j in range(len(roles)):
            if i!=j:
                userset1 = roles[i][0]
                userset2 = roles[j][0]
                # roles[j] is the parent role of roles[j]
                # roles[i] has higher priviledge than roles[j]
                if userset1.issubset(userset2):
                    hierarchyMatrix[i, j] = 1 

removes = list()
def dfsReduceRecursive(roles, i):
        global hierarchyMatrix, removes  
        numOfparentroles = np.sum(hierarchyMatrix[i])
        numOfchildroles = np.sum(hierarchyMatrix[:,i])
        # it means the roles[i] has more permission functions than some roles 
        # and no other role has more permission functions  than roles[i]
        if numOfparentroles > 0 and numOfchildroles == 0: 
            userSet1, funcSet1 = roles[i]
            for j in range(len(hierarchyMatrix[i])):
                if hierarchyMatrix[i, j] == 1:
                    parentrole = roles[j]
                    userSet2, funcSet2 = parentrole
                    funcSet1 = funcSet1.difference(funcSet2)
            # it means this role can be removed as their parentroles has contain all its functions .
            if len(funcSet1)==0:
                removes.append(i)
            else:
                # it means this role need be minimized to a new roles 
                roles[i] = (userSet1, funcSet1)  
            hierarchyMatrix[i] = np.zeros(len(hierarchyMatrix[i]))

        # it means the roles[i] has no more permission functions than any some roles 
        # but there exists other roles have more permission functions than roles[i]
        elif numOfparentroles == 0 and numOfchildroles >0:
            for j in range(len(hierarchyMatrix[:,i])):
                if hierarchyMatrix[j,i] == 1:
                    dfsReduceRecursive(roles, j)
        # it means the roles[i] has more permission functions than some roles 
        # and there exists other roles have more permission functions  than roles[i]
        elif numOfparentroles > 0 and numOfchildroles >0:
            userSet1, funcSet1 = roles[i]
            for j in range(len(hierarchyMatrix[i])):
                if hierarchyMatrix[i, j] == 1:
                    parentrole = roles[j]
                    userSet2, funcSet2 = parentrole
                    funcSet1 = funcSet1.difference(funcSet2)
            # it means this role can be removed as their parentroles has contain all its functions .
            # in this case, we need first to handle the role that has much more permission functions.
            for j in range(len(hierarchyMatrix[:,i])):
                if hierarchyMatrix[j,i] == 1:
                    dfsReduceRecursive(roles, j)
            
            # it means this role can be removed as their parentroles has contain all its functions .
            if len(funcSet1)==0:
                removes.append(i)
            else:
                # it means this role need be minimized to a new roles 
                roles[i] = (userSet1, funcSet1) 
            hierarchyMatrix[i] = np.zeros(len(hierarchyMatrix[i]))

def ReduceMain(roles):
    global hierarchyMatrix, removes
    hierarchyMatrix = None 
    removes = list() 
    buildRoleHierarchy(roles)
    for i in range(len(roles)):
        dfsReduceRecursive(roles, i)

    # print("removes: ", removes)
    removes = list(set(removes))
    removes.sort(reverse=True)
    [roles.pop(i) for i in removes]

    constant_roles = list()
    for role in roles:
        constant_roles.append((frozenset(role[0]), frozenset(role[1])))

    return constant_roles 

class GA_RM:
    def __init__(self, permissionMatrix, freqMatrix, simratio, userLabels, userMap, gene_encoding, address):
        self.userMap = userMap 
        self.userLabels = userLabels
        self.simratio = simratio
        self.df = permissionMatrix.T 
        self.permissions = list(permissionMatrix.columns)
        self.users = list(permissionMatrix.index)
        self.R = self.df.to_numpy()
        # U: users, A: funcs
        self.A, self.U = self.R.shape 
        print(f"No.user: {self.U}; No.func: {self.A}")
        self.start = time.time()       
        
        self.G_DF = self.df
     
        self.G_A = self.A
        self.G_U = self.U
        self.freqMatrix = freqMatrix
        self.gene_encoding = gene_encoding
        self.Debug = False
        self.address = address 

        self.verybasic_usergroups, self.UPA = self.createBasicLatticeRoles(self.users)
        self.basicRoles = ReduceMain(copy.copy(self.verybasic_usergroups))
        self.training_users = set()
        self.test_users = set()
    
    def translateLattice2Role(self, lattice):
        roles = []
        for function_set, user_set in lattice:
            if len(function_set) >0 and len(user_set)>0:
                roles.append((frozenset(user_set), frozenset([ func_str.replace("(","").replace(",)","").replace("'","") for func_str in function_set]))) 
        return roles  
    
    def createBasicLatticeRoles(self, users):
        df = self.G_DF.loc[:, list(map(int, users))]        
        # permissions
        objects = map(lambda id: str(id), df.index.tolist())
        # users
        properties = map(lambda id: str(id), df.columns.tolist())

        bools = list(df.fillna(False).astype(bool).itertuples(index=False, name=None))
        lattice = concepts.Context(objects, properties, bools)

        roles = self.translateLattice2Role(lattice.lattice)
        return roles, df.fillna(False).astype(bool)
      
    def getTrainingBasicRoles(self):
        return self.basicRoles, self.UPA

    def getTestingBasicRoles(self):
        return self.basicRoles, self.UPA
    
    def getUserFunctionCount(self, user, method):
        return self.freqMatrix.loc[int(user), method]
    
    @lru_cache(maxsize=None)
    def getAFV(self, role):
            user_set, func_set =  role
            # Consider global function usage by the users of candidate roles
            union_functions = list(set(self.permissions))
            # initialize average frequency vector for each candidate roles
            n  = len(union_functions)
            AFV = np.zeros(n)
            for i in range(n):
                function = union_functions[i]
                AFV[i] = np.sum(self.freqMatrix.loc[list(map(int, user_set)), function].to_numpy())/len(user_set)
            return AFV
    
    @lru_cache(maxsize=None)
    def calcsimilarity(self, roleA, roleB):
            AFV_a = self.getAFV(roleA)
            AFV_b = self.getAFV(roleB)
            similarity = 1 - distance.cosine(AFV_a, AFV_b)
            if self.Debug:
                print(f"AFV_a {AFV_a}")
                print(f"AFV_b {AFV_b}")
                print(f"similarity {similarity} for {roleA[1]} vs {roleB[1]}")
            return similarity
 
    def miningWithGAWith1DRealChromosome(self):
        Debug = False
        if self.U < 1:
            self.roles = []
            return self.roles 
        if self.U == 1:
            self.roles = []
            self.roles.append(([0], set(self.permissions)))
            return  self.roles

        trainBasicRoles, trainUPA = self.getTrainingBasicRoles()
        testBasicRoles, testUPA = self.getTestingBasicRoles()

        report = PrettyTable()
        report.title = "Basic roles statistics (id, len(users), functions)"
        report.field_names = ["RoleId", "Users", "Functions"]
        for index in range(len(trainBasicRoles)):
            role = trainBasicRoles[index]
            users, functions = len(role[0]), list(role[1])
            report.add_row([index, users, functions])
        print(report)
        
        # 1D real value Genome.
        # chromosome = [r1, r2, ..., rn]
        def translateChromosome2Roles(chromosome):
            badmergecount = 0
            roles = []
            new_chromosome = list(map(int, chromosome))
            mergeGroups = dict()
            for val in new_chromosome:
                mergeGroups[val] = np.where(np.array(new_chromosome)== val)[0]
            for groupkey, basicRoleIndices in mergeGroups.items():
                roles.append([trainBasicRoles[index] for index in basicRoleIndices])
            return roles, badmergecount
        
        def calculateTwoMineCompositeRolesSimilarity(crole1, crole2):
            sumSimError = 0
            for role1 in crole1:
                simError = 0
                for role2 in crole2:
                    simError = max(simError,  self.calcsimilarity(role1, role2))
                sumSimError += simError
            for role1 in crole2:
                simError = 0
                for role2 in crole1:
                    simError = max(simError,  self.calcsimilarity(role1, role2))
                sumSimError += simError

            return sumSimError/(len(crole1)+len(crole2))
    
        def calculateRoleSimilarityError(finalroles):
            simError = 0
            n = len(finalroles)
            if n==1:
                return simError
            for i in range(n-1):
                for j in range(i+1, n):
                    if i==j:
                        continue
                    simError = max(simError, calculateTwoMineCompositeRolesSimilarity(finalroles[i], finalroles[j]) )
            return simError
                
        def getClosestRoleInRoleSet(sourceRole, roles):
            maxSimilarity = 0
            closestRole = None
            for i in range(len(roles)):
                    roleA = roles[i]
                    roleA_functions = set(itertools.chain.from_iterable([ childrole[1] for childrole in roles[i]]))
                    roleB = sourceRole
                    similarity = len(roleA_functions.intersection(roleB[1]))/len(roleA_functions.union(roleB[1]))      
                    if similarity > maxSimilarity:
                        maxSimilarity  = similarity
                        closestRole = roleA
            return closestRole
        
        @lru_cache(maxsize=None)
        def getTestRoleL1Norm(role):
                users, funcs = role
                subTestUPA = testUPA.loc[:, list(map(int, users))]
                return np.count_nonzero(subTestUPA.to_numpy())
        
        @lru_cache(maxsize=None)
        def getDelta(predictedRole):
                users, funcs = predictedRole
                delta = len(users)*len(funcs) - np.count_nonzero(testUPA.loc[list(funcs), list(map(int, users))].to_numpy())
                return delta

        def calculateGeneralizationError(finalroles):
            predictedRoles = list()
            genError = 0
            counter = 0
            totalDelta = 0
            for role in testBasicRoles:
                closestRole = getClosestRoleInRoleSet(role, finalroles)
                if closestRole is None:
                    continue
                predictedRole =(role[0], frozenset(itertools.chain.from_iterable([childrole[1] for childrole in closestRole])))
                predictedRoles.append(predictedRole)
                testRoleL1Norm = getTestRoleL1Norm(role)
                delta = getDelta(predictedRole)
                totalDelta += delta
                error = delta/(testRoleL1Norm +  delta) 
                # error = delta/(len(role[0])*len(self.permissions)) 
                if error > 0:
                    genError += error 
                    counter += 1
            total_genError = 0
            if counter > 0:
                total_genError = genError/counter
            return total_genError
        
        @lru_cache(maxsize=None)
        def cached_eval_func(chromosome):
            score = 0.0
            finalroles, badmergecount = translateChromosome2Roles(chromosome)
            simErr = calculateRoleSimilarityError(finalroles)
            genErr = calculateGeneralizationError(finalroles)
            a = self.simratio
            b = 1 - a 
            score =  1 / (a * simErr + b * genErr + 0.001) 
            return score
            
        global eval_func
        def eval_func(chromosome):
            new_chromosome = list(map(int, chromosome))
            role_ids = sorted(list(set(new_chromosome)))
            new_chromosome = [ role_ids.index(_role_id) for _role_id in new_chromosome]
            # replace old chromosome with new chromosome
            for i in range(len(chromosome)):
                chromosome[i] = new_chromosome[i]
    
            return cached_eval_func(tuple(chromosome))

        def print_eval_func(chromosome):
            nonlocal Debug
            score = 0.0
            finalroles, badmergecount = translateChromosome2Roles(chromosome)
            simErr = calculateRoleSimilarityError(finalroles)
            genErr = calculateGeneralizationError(finalroles)
           
            a = self.simratio
            b = 1 - a 
            score = 1 / (a * simErr + b * genErr + 0.001) 
            finalroles = list(finalroles)
            
            returnroles = list()
            for role in finalroles:
                users = set(itertools.chain.from_iterable([ childrole[0] for childrole in role]))
                permissions = set(itertools.chain.from_iterable([ childrole[1] for childrole in role]))
                returnroles.append((users, permissions))
            print(chromosome)
            print(f"score: {score}, genErr: {genErr},  simErr: {simErr}, badmergecount: {badmergecount}")
            return score, genErr, simErr, returnroles, badmergecount 

        n = len(trainBasicRoles)
        assert n>1, "the number of basic roles must be greater than one"
        
        genome = G1DList.G1DList(n)
        genome.setParams(rangemin=0, rangemax=n)

        genome.initializator.set(Initializators.G1DListInitializatorReal)
        genome.mutator.set(Mutators.G1DListMutatorSwap)
        genome.mutator.add(Mutators.G1DListMutatorRealRange)
        genome.crossover.set(Crossovers.G1DListCrossoverSinglePoint) 
        genome.evaluator.set(eval_func)
          
        # ga = GSimpleGA.GSimpleGA(genome, seed=666)
        ga = GSimpleGA.GSimpleGA(genome, seed=2022)
        
        # Etilism ensures that best individuals will be carried to the next population (global optima)
        # Noted each new population would first be generated through selection, crossover and muation
        # Then the best indiv of previous generations will be compared to the current best indiv of the new population
        # Hereby, among them, the best indiv would be current etilism and will be put in the new population 
        # by replacing its best indiv
        ga.setElitism(True)
        ga.setElitismReplacement(1)
        ga.setSortType(Consts.sortType["scaled"])
        ga.selector.set(Selectors.GTournamentSelector)
        ga.internalPop.scaleMethod.set(Scaling.LinearScaling, weight=0.5)
        ga.internalPop.scaleMethod.add(Scaling.BoltzmannScaling, weight=0.1)
        ga.internalPop.scaleMethod.add(Scaling.ExponentialScaling, weight=0.1)
        ga.internalPop.scaleMethod.add(Scaling.PowerLawScaling, weight=0.1)
        ga.internalPop.scaleMethod.add(Scaling.SaturatedScaling, weight=0.1)
        ga.internalPop.scaleMethod.add(Scaling.SigmaTruncScaling, weight=0.1)

        ga.setMutationRate(0.10)
        ga.setCrossoverRate(0.99)
        ga.setPopulationSize(100)
        ga.setGenerations(200)
        
        ga.setMultiProcessing(True, full_copy=False, max_processes = 10)
        
        ga.evolve(freq_stats=10)
        bestindividual = ga.getPopulation().bestFitness()
        Debug = False
        score, genErr, simErr, bestRoles, badmergecount  = print_eval_func(bestindividual)
        Debug = False
        print(f"best role number: {len(bestRoles)}")
        no = 0
        for role in bestRoles:
            print(f"Role#{no}:{role[1]}")
            no += 1
        self.roles = bestRoles
        self.score, self.genErr, self.simErr, self.badMergeCount = (score, genErr, simErr, badmergecount)   
        return self.roles

    def process(self):
        roles = self.miningWithGAWith1DRealChromosome()
        newroles = list()
        for role in roles:
            roleuser = set([ self.userMap[int(user)] for user in role[0]]) 
            newroles.append((roleuser, role[1]))
        return newroles


def jaccard_func(set1, set2):
    try:
        return len(set(set1).intersection(set2))/len(set(set1).union(set2))
    except:
        return 0

def roleset_roleset_jaccard_func_old_donot_purse_one2one(roleset1, roleset2, t=1):
    def role_roleset_jaccard_func(role, roleset):
        max_jaccard_sim = 0
        matched_role = None
        for _role in roleset:
            jaccard_sim = jaccard_func(role, _role)
            if jaccard_sim > max_jaccard_sim:
                max_jaccard_sim = jaccard_sim
                matched_role = _role
        return  max_jaccard_sim, matched_role
    
    roleset1 = copy.copy(roleset1)
    roleset2 = copy.copy(roleset2)
    n1 = len(roleset1)
    n2 = len(roleset2)
    if n1 == 0 or n2 == 0:
        return 1
    if n1>n2:
        roleset_tmp = roleset1
        roleset1 = roleset2
        roleset2 = roleset_tmp
    
    sum_sim = 0
    matched_roles = list()
    counter = 0
    for role in roleset1:
        maxjaccardsim, matched_role = role_roleset_jaccard_func(role, roleset2)
        if matched_role is not None:
            matched_roles.append(matched_role)
        sum_sim = sum_sim + maxjaccardsim
        counter += 1
 
    for matched_role in matched_roles:
        if roleset2.count(matched_role)>0:
            roleset2.remove(matched_role)

    for restrole in roleset2:
        maxjaccardsim, _ = role_roleset_jaccard_func(restrole, roleset1)
        if maxjaccardsim > t:
            sum_sim = sum_sim + maxjaccardsim
            counter += 1

    if counter == 0:
        return -1
    return sum_sim/counter

def compareRoleSets(mined_roles, deployed_roles, t=1.0):
    print("**********************")
    print(mined_roles, deployed_roles)
    permissionless_functions = set(itertools.chain.from_iterable(mined_roles)).difference(\
        set(itertools.chain.from_iterable(deployed_roles)))
    deployed_roles.append(permissionless_functions)
    rolesroles_sim1 = roleset_roleset_jaccard_func_old_donot_purse_one2one(mined_roles, deployed_roles, t=t)
    deployed_roles.remove(permissionless_functions)
    mined_roles_remove_permissionless = [ role.difference(permissionless_functions) \
        for role in mined_roles if len(role.difference(permissionless_functions))>0]

    deployed_roles_remove_permissionless = [ role.difference(permissionless_functions) \
        for role in deployed_roles if len(role.difference(permissionless_functions))>0]
                 
    rolesroles_sim2 = roleset_roleset_jaccard_func_old_donot_purse_one2one(mined_roles_remove_permissionless, \
        deployed_roles_remove_permissionless, t=t)
    print(rolesroles_sim1, rolesroles_sim2)
   
    return  rolesroles_sim1, rolesroles_sim2


def label_func(role_permission_functions, other_permission_functions, ABIfunctions):
    rp_funcs = role_permission_functions
    op_funcs = other_permission_functions
    abi_funcs = ABIfunctions
    
    def func(observed_methods, mined_roles):
        # label mined roles with ground truth, namely rp_funcs+op_funcs
        # using jaccard similarity
        # get all relevant roles
        deployed_role_permissionset = set()
        deployed_role_labelset = set()

        relevant_roles = set(map(lambda rp: rp[1], rp_funcs)).union(set(map(lambda op: op[1], op_funcs)))
        relevant_role_permissions = dict()
        permissioned_functions = set()
        for role in relevant_roles:
            relevant_role_permissions[role] = {}
            relevant_role_permissions[role]["orig"] = set(filter(lambda rp: rp[1]==role, rp_funcs)).union(set(filter(lambda op: op[1]==role, op_funcs)))
            relevant_role_permissions[role]["simple"] =set(map(lambda p: p[0], relevant_role_permissions[role]["orig"]))
            permissioned_functions.update(relevant_role_permissions[role]["simple"])
        
        permissionless_functions = set(ABIfunctions).difference(permissioned_functions)
            
        results = list()
        for role in mined_roles:
            candidate_relevant_roles = list(filter(lambda candidate: \
                len(relevant_role_permissions[candidate]["simple"].intersection(role))>0, \
                relevant_roles))
            if len(candidate_relevant_roles)==0:
                # normal users which call only permissionless functions
                results.append(((NORMAL_USER), role, jaccard_func(role, permissionless_functions)))
                continue 
            jaccard_max_sim = 0
            label = None 
            for i in range(1, len(candidate_relevant_roles)+1):
                for role_label_pair in combinations(candidate_relevant_roles, i):
                    # calculate the jaccard similarity
                    role_label_pair_functions = set(itertools.chain.from_iterable(\
                        [ relevant_role_permissions[r]["simple"] for r in role_label_pair]))
                    # print(set(role).intersection(permissioned_functions), role_label_pair_functions)
                    if set(role).intersection(permissioned_functions).issubset(role_label_pair_functions):
                        jaccard_sim = jaccard_func(set(role).intersection(permissioned_functions), role_label_pair_functions)
                    else:
                        jaccard_sim = 0
                    if jaccard_sim > jaccard_max_sim:
                        jaccard_max_sim = jaccard_sim
                        label = role_label_pair 
                        deployed_role_labelset.update(role_label_pair)                       
            results.append((label, role, jaccard_max_sim))
      
        deployed_role_permissionset =  [ ((role_label), relevant_role_permissions[role_label]["simple"].intersection(observed_methods)) for role_label in deployed_role_labelset]
        all_role_permissionset = [ ((role_label), relevant_role_permissions[role_label]["simple"]) for role_label in relevant_roles]
        return deployed_role_labelset, deployed_role_permissionset, all_role_permissionset, results 
        
    return func

def getSetOfSimilarityMetrics(mined_roles, deployed_roles):
    role_sim10_1, role_sim10_2 = compareRoleSets(mined_roles=mined_roles, deployed_roles=deployed_roles, t=1.0)

    role_sim05_1, role_sim05_2 = compareRoleSets(mined_roles=mined_roles, deployed_roles=deployed_roles, t=0.5)
            
    role_sim025_1, role_sim025_2 = compareRoleSets(mined_roles=mined_roles, deployed_roles=deployed_roles, t=0.25)

    role_sim00_1, role_sim00_2 = compareRoleSets(mined_roles=mined_roles, deployed_roles=deployed_roles, t=0)
    
    return role_sim10_1, role_sim10_2, role_sim05_1, role_sim05_2, role_sim025_1, role_sim025_2, role_sim00_1, role_sim00_2

def initExcelHead(ws_result):
    ws_result.title = "RoleMiningResult"
    ws_result["A1"] = "(Alpha, Beta)"
    ws_result["B1"] = "Address"
    ws_result["C1"] = "Time"
    ws_result["D1"] = "RoleNumber"
    ws_result["E1"] = "Roles"  
    ws_result["F1"] = "LabeledRoles"  
    ws_result["G1"] = "DeployedRoles"  
   
    ws_result[f"H{1}"] = "Role Number Ratio"
    ws_result[f"I{1}"] = "t=1"
    ws_result[f"J{1}"] = "t=0.5"
    ws_result[f"K{1}"] = "t=0.25"
    ws_result[f"L{1}"] = "t=0"

    ws_result[f"M{1}"] = "t=1"
    ws_result[f"N{1}"] = "t=0.5"
    ws_result[f"O{1}"] = "t=0.25"
    ws_result[f"P{1}"] = "t=0"

def appendExcelRow(ws, n, simratio, address, timecost,\
     roleNumber, mined_roles, labledroles, groundtruth_roles, number_ratio, \
         role_sim10_1, role_sim10_2, role_sim05_1, role_sim05_2, \
             role_sim025_1, role_sim025_2, role_sim00_1, role_sim00_2):
        ws[f"A{n}"] = f"({simratio}, {1 - simratio})"
        ws[f"B{n}"] = address
        ws[f"C{n}"] = timecost
        ws[f"D{n}"] = roleNumber
        ws[f"E{n}"] = str(mined_roles) 
        ws[f"F{n}"] = str(labledroles)
        ws[f"G{n}"] = str(groundtruth_roles)
        ws[f"H{n}"] = str(number_ratio)
        ws[f"I{n}"] = role_sim10_1 
        ws[f"J{n}"] = role_sim05_1 
        ws[f"K{n}"] = role_sim025_1 
        ws[f"L{n}"] = role_sim00_1
        ws[f"M{n}"] = role_sim10_2 
        ws[f"N{n}"] = role_sim05_2 
        ws[f"O{n}"] = role_sim025_2 
        ws[f"P{n}"] = role_sim00_2

def _main_(args):
    global permissionless_functions
    simratio = args.simratio
    workdir = args.benchmark
    xlsx = args.groundtruth
    wb = openpyxl.load_workbook(xlsx, read_only=False)    
    ws = wb.active 
    contracts_limited_number = args.limit
    rows = list(ws.rows)[1:]
    counts = list()
    wb_result = Workbook()
    ws_result = wb_result.active
    initExcelHead(ws_result=ws_result)
    index = 2
    for row in rows:
        try:
            if contracts_limited_number == 0:
                break 
            address = row[0].value
            if address is None: 
                break 
            contractName = os.path.splitext(os.path.basename(row[0].hyperlink.target))[0]

            print(f"{contractName}: ", address)
            # abi_file = os.path.join(os.path.dirname(row[0].hyperlink.target),contractName+".abi")
            abi_file = os.path.join(workdir, address, contractName+".abi")
            ABIfunctions = getABIfunctions(abi_file)

            with open(os.path.join(workdir, address, "user_all.json"), "r") as f:
                    usercall_statistics = json.load(f)
                    userCount = len(usercall_statistics["data"]["ethereum"]["smartContractCalls"])
                    print("userCount:", userCount)     
                    counts.append(userCount)
                    if userCount > 100:
                        if row[3].value!="" and row[3].value!="set()":
                            role_permission_functions = eval(row[3].value)
                        else:
                            role_permission_functions = set()
                        if row[4].value!="" and row[4].value!="set()":
                            other_permission_functions = eval(row[4].value)
                        else:
                            other_permission_functions = set()

                        print(role_permission_functions, other_permission_functions)
                        eval_func = label_func(role_permission_functions, other_permission_functions, ABIfunctions)
                        
                        start = time.time()
                        observed_methods, verybasic_userGroups, mined_roles = lightweightrolemining(address=address, gene_encoding="real", simratio=simratio, workdir=workdir)
                        end = time.time()
                        
                        mined_roles = [role[1] for role in mined_roles ]
                        
                        _, _, _, labeled_mined_roles = eval_func(observed_methods, mined_roles)
                        verybasic_roles = [ usergroup[1] for usergroup in verybasic_userGroups]
                        deployed_role_labelset, deployed_role_permissionset, all_role_permissionset, _ = eval_func(observed_methods, verybasic_roles)

                        deployed_roles = [role[1] for role in deployed_role_permissionset]
                      
                        number_ratio = len(mined_roles)/(len(deployed_roles)+1)
                        # rolesroles_sim1, rolesroles_sim2 = compareRoleSets(mined_roles=mined_roles, deployed_roles=deployed_roles)
                        role_sim10_1, role_sim10_2, role_sim05_1, role_sim05_2, \
                            role_sim025_1, role_sim025_2, role_sim00_1, role_sim00_2 \
                                = getSetOfSimilarityMetrics(mined_roles=mined_roles, deployed_roles=deployed_roles)
                         
                        appendExcelRow(ws_result, index, simratio, address, end-start, len(mined_roles), mined_roles, labeled_mined_roles, deployed_role_permissionset, 
                        number_ratio, role_sim10_1, role_sim10_2, role_sim05_1, role_sim05_2, \
                            role_sim025_1, role_sim025_2, role_sim00_1, role_sim00_2
                        )
                        contracts_limited_number -= 1
            index += 1
        except:
            traceback.print_exc()
            continue  
    wb_result.save(args.output)

def main():
    start = time.time()
    parser = argparse.ArgumentParser(description='SPCONMiner, Mining smart contract role structures')
       
    parser.add_argument('--benchmark', type=str, default = "./ISSTA2022/RoleMiningBenchmarkandResults/OpenZeppelin1000calls10methods" , 
                        help='benchmark directory (default ./ISSTA2022/RoleMiningBenchmarkandResults/OpenZeppelin1000calls10methods)')

    parser.add_argument('--groundtruth', type=str, default = "./ISSTA2022/RoleMiningBenchmarkandResults/OpenZeppelin1000calls10methods-label.xlsx" , 
                        help='the labelled role structure (ground truth) for the benchmark (default ./ISSTA2022/RoleMiningBenchmarkandResults/OpenZeppelin1000calls10methods-label.xlsx)')
    
    parser.add_argument('--output', type=str, default="./result-OpenZeppelin_spconminer.xlsx",
                        help= "the output file containing result of mined role structure and its comparison with the ground truth (./result-OpenZeppelin_spconminer.xlsx)")
    
    parser.add_argument('--simratio', type=float, default = 0.40, 
                        help='ratio of simErr for GA. (default 0.40)')
    
    parser.add_argument('--limit', type=int, default = 50, 
                        help='how many benchmark contracts are inspected for the role mining. (default 50)')
    args = parser.parse_args()
    assert args is not None 
    try:
        _main_(args)
    except:
        traceback.print_exc()
        pass 
    print(f"total timecost: {time.time() - start} seconds")

if __name__ == "__main__":
    main()

