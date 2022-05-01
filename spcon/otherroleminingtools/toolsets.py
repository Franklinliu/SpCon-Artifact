from networkx.algorithms.operators.product import rooted_product
import numpy as np
from numpy.lib.function_base import append 
import pandas as pd 
import os 
import sys
import copy 
import networkx as nx 
import concepts
import matplotlib.pyplot as plt
import graphviz
from json import JSONEncoder
import json 
import pickle
import jsonpickle
import jsonpickle.handlers
import time
import math

from scipy.spatial import distance 
from sklearn.metrics import mean_squared_error
import json
import time
import concepts
from scipy.spatial import distance
import numpy as np
import pandas as pd
import os 
import sys
from openpyxl import Workbook
import openpyxl
from itertools import combinations
import itertools 
from functools import lru_cache
import requests
import traceback
import timeout_decorator
from web3 import Web3
NORMAL_USER = ("NORMAL_USER")

@lru_cache(maxsize=None)
def getMethodName(hex_signature):
    try:
        headers = {'X-API-KEY': 'BQYEtWs7QzCdCwkJKVhmTDp3RFTWAUEP'}
        response = requests.get(f'https://www.4byte.directory/api/v1/signatures/?hex_signature={hex_signature}',
                                          headers=headers)
        body = response.content
        methodName = json.loads(body.decode("utf8"))["results"][0]["text_signature"].split("(")[0]
        return methodName
    except:
        return hex_signature
def getABIfunctions(abi_file):
    ABIfunctions = list()
    abis = json.load(open(abi_file))
    for abi in abis:
        if abi["type"] == "function" and abi["stateMutability"]!="view":
            ABIfunctions.append(abi["name"])
    return ABIfunctions 

def getABI_file(directory):
    items = os.listdir(directory)
    abi_file = None 
    for item in items:
        if item.endswith("abi"):
            abi_file = os.path.join(directory, item)
            break 
    if abi_file:
        return abi_file
    raise Exception("ABI file not found")        

def getABIfunction_signature_mapping(abi_file):
    mappings = dict()
    abi = json.load(open(abi_file))
    signatures = [ (func['name'], '{}({})'.format(func['name'],','.join([input['type'] for input in func['inputs']]))) for func in [obj for obj in abi if obj['type'] == 'function']]
    names = set()
    for sig in signatures:
        name, full_name = sig 
        bytes4_sig = Web3.sha3(text=full_name)[0:4].hex()
        mappings[bytes4_sig] = sig
        names.add(name)

    def func(bytes4_sig):
        if bytes4_sig in names:
            return bytes4_sig
        if not bytes4_sig.startswith("0x"):
            bytes4_sig = "0x"+bytes4_sig
        if bytes4_sig in mappings:
            return mappings[bytes4_sig][0]
        else:
            return getMethodName(bytes4_sig)
    return func 
    


def lightweightrolemining(address, workdir):
    smartcontractCalls = list()
    try:
        bytes4_mapping_func = getABIfunction_signature_mapping(getABI_file(os.path.join(workdir, address)))
        print("loaded abi.")
    except:
        bytes4_mapping_func = getMethodName
        pass 
        
    with open(os.path.join(workdir, address, "user_all2.json"), "r") as f:
        usercall_statistics = json.load(f)
    for usercall in usercall_statistics["data"]["ethereum"]["smartContractCalls"]:
        caller, count, function, success = usercall["caller"]["address"], usercall["count"], usercall["smartContractMethod"]["name"] if usercall["smartContractMethod"]["name"]!=None else  bytes4_mapping_func(usercall["smartContractMethod"]["signatureHash"]),  usercall["success"]
        smartcontractCalls.append((caller, function, count, success))
    def getIdCounter():
            hashMap = dict()
            userMap = dict()
            def _getId(user):
                user = user.lower()
                if user in hashMap:
                    return hashMap[user]
                else:
                    hashMap[user] = len(userMap)
                    userMap[len(userMap)] = user
                    return hashMap[user]
           
            return userMap, _getId 
    UF = list()
    userFunctions = dict()
    functions = set()
    userMap, IdCounter = getIdCounter()
                        
    for smartcontractCall in smartcontractCalls: 
        user, function, count, success = smartcontractCall
        if function!="Contract Creation" and (success == 1 or success is True):
                functions.add(function)
                if IdCounter(user) not in userFunctions:
                    userFunctions[IdCounter(user)] = list()
                userFunctions[IdCounter(user)].append(function)
                UF.append((IdCounter(user), function, count))
    functions = list(functions)
    title = functions
    matrix = [ [1 if function in set(userFunctions[user_id]) else 0 for function in functions]  for user_id in userFunctions ]
    permissionMatrix =  pd.DataFrame(np.array(matrix), columns=title)
    print(len(functions), " functions", functions)
        
    roleminer = GA_RM(permissionMatrix)      
    return functions, roleminer.verybasic_usergroups, permissionMatrix


class GA_RM:
    def __init__(self, permissionMatrix):
        # shufle the users for later sampling
        #jjjself.df = permissionMatrix.sample(n = permissionMatrix.shape[0])
        self.df = permissionMatrix.T 
        self.permissions = list(permissionMatrix.columns)
        self.users = list(permissionMatrix.index)
        self.R = self.df.to_numpy()
        # U: users, A: funcs
        self.A, self.U = self.R.shape 
        print(f"No.user: {self.U}; No.func: {self.A}")
        self.start = time.time()       
        
        self.G_DF = self.df
     
        self.G_A = self.A
        self.G_U = self.U

        self.verybasic_usergroups, _ = self.createBasicLatticeRoles(self.users)
    
    def translateLattice2Role(self, lattice):
        roles = []
        for function_set, user_set in lattice:
            if len(function_set) >0 and len(user_set)>0:
                roles.append((set(user_set), set([ func_str.replace("(","").replace(",)","").replace("'","") for func_str in function_set]))) 
        return roles  
    
    def createBasicLatticeRoles(self, users):
        df = self.G_DF.loc[:, list(map(lambda u: int(u), users))]        
        # permissions
        objects = map(lambda id: str(id), df.index.tolist())
        # users
        properties = map(lambda id: str(id), df.columns.tolist())

        bools = list(df.fillna(False).astype(bool).itertuples(index=False, name=None))
        lattice = concepts.Context(objects, properties, bools)

        roles = self.translateLattice2Role(lattice.lattice)
        return roles, df.fillna(False).astype(bool)


class NodeHandler(jsonpickle.handlers.ArrayHandler):
    def flatten(self, obj, data):
            data["intent"] = obj.intent 
            data["extent"] = obj.extent
    def restore(self, obj, data):
        obj.indent = data["intent"]
        obj.extent = data["extent"]

class NodeEncoder(JSONEncoder):
    def default(self, object):
        if isinstance(object, Node):
            # print(object.__dict__)
            return {"extent": object.extent, "intent": object.intent}
        else:
            # call base class implementation which takes care of
            # raising exceptions for unsupported types
            return json.JSONEncoder.default(self, object)

class Node(object):
    def __init__(self, concept, intent, extent, isEmptyConcept=False):
        global counter
        self.concept = concept
        self.intent = intent
        self.extent = extent
        self.isEmptyConcept = isEmptyConcept

    def __str__(self):
        # return self.__dict__
        if self.isEmptyConcept == False:
            return "{0}\n---------\n{1}".format(str(self.intent), len(self.extent))
        else:
            return "{0}\n---------\n{1}".format("{}", len(self.extent))
    
    def __eq__(self,other):
        if other is None:
            return False
        
        return self.concept == other.concept or (self.intent == other.intent and self.extent == other.extent and self.isEmptyConcept ==other.isEmptyConcept)
    
    def __hash__(self):
        return hash(self.concept)

# HierarchyMiner
class HMLattice:
    def __init__(self, permissionMatrix, starttime):
        self.start  = starttime
        self.df = permissionMatrix
        self.permissions = permissionMatrix.columns
        self.R = permissionMatrix.to_numpy()
        # self.U: number of object set, self.A: number of properties set
        self.U, self.A = self.R.shape 
        # Lattice
        self.L =  list()
        print(self.U, self.A)

    def graphviz(self, G, rootNode, filename, directory=None, render = True, view = False,
            **kwargs):
            """Return graphviz source for visualizing the lattice graph."""
            dot = graphviz.Digraph(name=self.__class__.__name__,
                                comment=repr(G),
                                filename=filename, directory=directory,
                                node_attr={'style':"rounded",  'shape':'record', 'label': ''},
                                edge_attr={'dir': 'none'},
                                **kwargs)
            if rootNode is None or G is None:
                return 
            LIMIT = 6
            self.no = 0
            def node_name(node):
                # self.no += 1
                return "c"+str(hash(node)) 
            self.drawed = set()
            def drawNode(node):
                name = node_name(node)
                if not (name in self.drawed):
                    if node.extent:
                      
                        extent = "{"+",".join([ func_str.replace("(","").replace(",)","").replace("'","") for func_str in  node.extent])+"}"
                    else:
                        extent = "{EmptySet}"
                    if node.intent:
                        if len(node.intent)>LIMIT:
                            intent ="{"+",".join([ "user"+str(user_id) for user_id in node.intent[:LIMIT]])+"..."+str(len(node.intent))+" users}"
                        else:
                            intent="{"+",".join([ "user"+str(user_id) for user_id in node.intent])+"}"
                    else:
                        intent = "{EmptySet}"
                    dot.node(name, label="{"+f"{extent}|{intent}"+"}")
                    self.drawed.add(name)
                else:
                    return
               
                dot.edges((name, node_name(c)) for c in G.successors(node))
                # for c in G.successors(node):
                #     drawNode(c)
            # drawNode(rootNode)
            for node in G.nodes:
                drawNode(node)
            if render or view:
                dot.render(view=view)
            return dot 


    def lattice_graph(self):
        def Sen(r):
                return len(list(self.G.successors(r)))

        def Jun(r):
                return len(list(self.G.predecessors(r)))

        # Thr(r) is the set of pairs of roles (ri,rj) such that, without role r, ri would no longer be senior to rj
        def Thr(r): 
                successors = list(self.G.successors(r))
                predecessors = list(self.G.predecessors(r))
                
                connectivity_pairs_old = set()
                
                [ [connectivity_pairs_old.add((startnode, endnode)) for endnode in endnodes] for startnode, endnodes  in nx.algorithms.all_pairs_node_connectivity(self.G).items()]

                self.G.remove_node(r)

                connectivity_pairs_new = set()
                
                [ [connectivity_pairs_new.add((startnode, endnode)) for endnode in endnodes] for startnode, endnodes  in nx.algorithms.all_pairs_node_connectivity(self.G).items()]


                thr =  len(connectivity_pairs_new.difference(connectivity_pairs_old)) - len(successors) - len(predecessors)

                self.G.add_node(r)
                for successor in successors:
                    self.G.add_edge(r, successor)
                
                for predecessor in predecessors:
                    self.G.add_edge(predecessor, r)

                return thr

        def creatG():
            NodeDict = dict()
            root = Node(self.c.lattice.supremum, self.c.lattice.supremum.intent, self.c.lattice.supremum.extent, isEmptyConcept=True)
            NodeDict[hash(self.c.lattice.supremum)] = root
            self.G = nx.DiGraph()
            def construct(root):
                for subconcept in root.concept.lower_neighbors:
                    if hash(subconcept) not in NodeDict:
                        NodeDict[hash(subconcept)] = Node(subconcept, subconcept.intent, subconcept.extent)
                    subnode = NodeDict[hash(subconcept)]
                    self.G.add_edge(root, subnode)
                    construct(subnode)
                return root

            def dfsReduceG(root):
                successors = self.G.successors(root)
                for successor in successors:
                    # intent is users
                    # extent is functions
                    root.extent = tuple(set(root.extent).difference(set(successor.extent)))
                    dfsReduceG(successor)
                    successor.intent = tuple(set(successor.intent).difference(set(root.intent)))

            construct(root)
            dfsReduceG(root)
            return root, NodeDict, self.G

        root, NodeDict, self.G = creatG()

        Wr, Wu, Wp, Wh, Wd = (1,1,1,1,1)

        for h, node in NodeDict.items():
            if not self.G.has_node(node):
                continue
            # print(Sen(node), Jun(node), Thr(node))
            if len(node.extent) == 0 and len(node.intent) == 0:
                if Wh*(Sen(node) + Jun(node))+Wr >= Wh*Thr(node):
                    predecessors = self.G.predecessors(node)
                    successors =  self.G.successors(node)
                    self.G.remove_node(node)
                    for predecessor in predecessors:
                        for successor in successors:
                            if successor not in self.G.successors(predecessor):
                                self.G.add_edge(predecessor, successor)
                    
            elif len(node.extent)!=0 and len(node.intent)==0:
                if Wr + Wu * self.U + Wh * (Sen(node) + Jun(node)) >= Wu * self.U * Jun(node) + Wh * Thr(node):

                    predecessors = self.G.predecessors(node)
                    successors = self.G.successors(node)
                    
                    self.G.remove_node(node)
                    
                    for predecessor in predecessors:
                        for successor in successors:
                            if successor not in self.G.successors(predecessor):
                                self.G.add_edge(predecessor, successor)
                    
                    for successor in successors:
                        successor.intent = tuple(set(list(node.intent)).union(set(list(successor.intent))))


            elif len(node.extent) == 0 and len(node.intent)>0:
                if Wr + Wp * self.A + Wh * (Sen(node) + Jun(node)) >= Wp * self.A * Sen(node) + Wh * Thr(node):
                    predecessors = self.G.predecessors(node)
                    successors = self.G.successors(node)
                  
                    self.G.remove_node(node)
                       
                    for predecessor in predecessors:
                        for successor in successors:
                            if successor not in self.G.successors(predecessor):
                                self.G.add_edge(predecessor, successor)
                    
                    for predecessor in predecessors:
                        predecessor.extent = tuple(set(list(node.extent)).union(set(list(predecessor.extent))))
        
        def dfsRecoverG(child):
                precessors = self.G.predecessors(child)
                for precessor in precessors:
                    # intent is users
                    # extent is functions
                    dfsRecoverG(precessor)
                    child.intent = tuple(set(child.intent).union(set(precessor.intent)))
                    precessor.extent = tuple(set(child.extent).union(set(precessor.extent)))
                return child
        def recoverG():
            for node in self.G.nodes:
                dfsRecoverG(node)
        self.root = root
        recoverG()
        return
        
    @timeout_decorator.timeout(5*60)
    def process(self):
        
        objects = map(lambda id: str(id),self.df.index.tolist())
        properties = list(self.df)
        bools = list(self.df.fillna(False).astype(bool).itertuples(index=False, name=None))
    
        c = concepts.Context(objects, properties, bools)
        self.c = c        

        self.lattice_graph()

        roles = [ set(node.intent) for node in self.G.nodes ]
        return roles


# HP Role Minimization
class HPr:
    def __init__(self, permissionMatrix, starttime):
        self.start = starttime
        self.permissions = list(permissionMatrix.columns)
        self.userPermission = permissionMatrix.to_numpy()
        self.usernumber, self.permissionnumber = self.userPermission.shape 
        self.roles = list()
        print(self.usernumber, self.permissionnumber)
    
    def initialize(self):
        self.coveredpermission = set()
        self.covereduser = set()
    
    @timeout_decorator.timeout(20*60)
    def process(self):
        
        self.initialize()
        self.strategy_chooseuser_fewestuncoveredpermissions()
        self.strategy_choosepermission_fewestuncoveredusers()
        roles = [set(role[1]) for role in self.roles]
        return roles  

    def strategy_choosepermission_fewestuncoveredusers(self):
        while True:
            permissionUserCount = None 
            for i in range(len(self.permissions)):
                if self.permissions[i] not in self.coveredpermission:
                    if permissionUserCount is None:
                        permissionUserCount = np.sum(self.userPermission[:,])
                    else:
                        permissionUserCount = min(permissionUserCount, np.sum(self.userPermission[:,]))
            if permissionUserCount is None:
                break
            for i in range(len(self.permissions)):
                if self.permissions[i] not in self.coveredpermission:
                    if permissionUserCount == np.sum(self.userPermission[:,]):
                        role = (self.Up(i),self.Pp(i))
                        self.roles.append(role)
                        for u in role[0]:
                            self.covereduser.add(u)
                        for p in role[1]:
                            self.coveredpermission.add(p)
                        break

    def strategy_chooseuser_fewestuncoveredpermissions(self):
        while True:
            userpermissionCount = None 
            for user in range(self.usernumber):
                if user not in self.covereduser:
                    if userpermissionCount is None:
                        userpermissionCount = np.sum(self.userPermission[user])
                    else:
                        userpermissionCount = min(userpermissionCount, np.sum(self.userPermission[user]))
            if userpermissionCount is None:
                break
            for user in range(self.usernumber):
                if user not in self.covereduser:
                    if userpermissionCount == np.sum(self.userPermission[user]):
                        role = (self.Uu(user),self.Pu(user))
                        self.roles.append(role)
                        for u in role[0]:
                            self.covereduser.add(u)
                        for p in role[1]:
                            self.coveredpermission.add(p)
                        break
    # get permission sets of user#u                    
    def Pu(self, u):
        permissions = set()
        for i in range(len(self.userPermission[u])):
            if self.userPermission[u][i]==1:
                permissions.add(self.permissions[i])
        return permissions

    # get all users who have all of user#u's permissions
    def Uu(self, u):
        users = set()
        permissions = self.Pu(u)
        assert len(permissions)>0, "permission set is empty"
        for user in range(self.usernumber):
            user_permissions = set()
            [user_permissions.add(self.permissions[i]) if self.userPermission[user][i]==1 else "" for i in range(len(self.userPermission[user]))]
            # test if user_permissions are subset of permissions
            if len(user_permissions.intersection(permissions)) == len(permissions):
                users.add(user)
        return users
    
    # get all permissions assigned to all users in Up(p)
    def Pp(self, p):
        permissions = set()
        users = self.Up(p)
        assert len(users)>0, "users set is empty"
        for user in users:
            user_permissions = set()
            [user_permissions.add(self.permissions[i]) if self.userPermission[user][i]==1 else "" for i in range(len(self.userPermission[user]))]
            permissions = permissions.union(user_permissions)
        return permissions
    # get all users who have permission#p
    def Up(self, p):
        users = set()
        for i in range(len(self.userPermission[:,p])):
            if self.userPermission[i][p]==1:
                users.add(i)
        return users 
    

class ORCA:
    def __init__(self, permissionMatrix, starttime):
        self.Clusters = list()
        self.exclusiveClusterPair = list()
        self.outerClusters = list()
        self.PartialOrderOfClusters = set()
        self.permissions = list(permissionMatrix.columns)
        self.userPermission = permissionMatrix.to_numpy()
        self.start = starttime
        # print(self.permissions)
    def getPermissionIndex(self, permission):
        for i in range(len(self.permissions)):
            if self.permissions[i] == permission:
                return i
        raise permission+" is out of permission range!"
    def members(self, cluster):
        if isinstance(cluster, set):
            mbs = None
            for item in cluster:
                tmp = self.members(item)
                if mbs is None:
                    mbs = tmp 
                else:
                    mbs = mbs.intersection(tmp)
            # print(cluster if mbs is None else "")
            assert mbs is not None, self.Clusters 
            return mbs
        else:
            mbs = set()
            permission = cluster 
            usersOfPermission = self.userPermission[:,self.getPermissionIndex(permission)]
            for user in range(len(usersOfPermission)):
                if usersOfPermission[user] == 1:
                    mbs.add(user)
            
            return mbs 

    def rights(self, cluster):
        if isinstance(cluster, set):
            rts = None
            for item in cluster:
                tmp = self.rights(item)
                if rts is None:
                    rts = tmp 
                else:
                    rts = rts.union(tmp)
            return rts
        else:
            permission = cluster
            rts = set()
            rts.add(permission)
            return rts  

    def newcluster(self, cluster_a, cluster_b):
        cluster = set()
        if isinstance(cluster_a, set):
            cluster = cluster.union(cluster_a)
        else:
            cluster.add(cluster_a)
        if isinstance(cluster_b, set):
            cluster = cluster.union(cluster_b)
        else:
            cluster.add(cluster_b)
        return cluster

    def less(self, cluster_a, cluster_b):
        if isinstance(cluster_b, set) and isinstance(cluster_a, str):
            return cluster_a in cluster_b
        elif isinstance(cluster_a, set) and isinstance(cluster_b, set):
            for item in cluster_a:
                if item not in cluster_b:
                    return False
            return len(cluster_a) <= len(cluster_b)
        return cluster_a!=cluster_b

    def initialize(self):
        [self.Clusters.append(set([permission])) for permission in self.permissions]
        # print(self.Clusters)
        for cluster_a in self.Clusters:
                for cluster_b in self.Clusters:
                    assert cluster_a is not None
                    assert cluster_b is not None
    
    def mining(self):
        while True:
            m = 0
            for cluster_a in self.Clusters:
                for cluster_b in self.Clusters:
                    if not ( cluster_a ==  cluster_b or cluster_a.union(cluster_b) in self.exclusiveClusterPair or self.less(cluster_a, cluster_b) or self.less(cluster_b, cluster_a)):
                        m = max(m, len(self.members(cluster_a).intersection(self.members(cluster_b))))
            if m == 0:
                break
            r = 0
            for cluster_a in self.Clusters:
                for cluster_b in self.Clusters:
                    if not ( cluster_a ==  cluster_b or cluster_a.union(cluster_b) in self.exclusiveClusterPair or self.less(cluster_a, cluster_b) or self.less(cluster_b, cluster_a)):
                        if m ==  len(self.members(cluster_a).intersection(self.members(cluster_b))):
                            r = max(r, len(self.rights(cluster_a).intersection(self.rights(cluster_b))))
            for cluster_a in self.Clusters:
                for cluster_b in self.Clusters:
                    if not ( cluster_a ==  cluster_b or cluster_a.union(cluster_b) in self.exclusiveClusterPair or self.less(cluster_a, cluster_b) or self.less(cluster_b, cluster_a)):
                        if m ==  len(self.members(cluster_a).intersection(self.members(cluster_b))) and r == len(self.rights(cluster_a).intersection(self.rights(cluster_b))):
                            # print(cluster_a, cluster_b)
                            cluster = self.newcluster(cluster_a, cluster_b)
                            # if True:
                            #     # print(cluster_a, cluster_b)
                            #     print(cluster)
                            self.Clusters.append(cluster)
                            self.Clusters.remove(cluster_a)
                            self.Clusters.remove(cluster_b)
                            self.exclusiveClusterPair.append(cluster_a.union(cluster_b))
                           
                            self.outerClusters.append(cluster_a)
                            self.outerClusters.append(cluster_b)
                            
                            break
        return 
    
    @timeout_decorator.timeout(20*60)
    def process(self):
        self.initialize()
        self.mining() 
        roles = [ set(cluster) for cluster in self.Clusters]
        roles.extend( [ set(cluster) for cluster in self.outerClusters])
        return roles 


class GO:
    def __init__(self, permissionMatrix, starttime):
        self.Clusters = list()
        self.exclusiveClusterPair = list()
        self.removeClusters = list()
        self.PartialOrderOfClusters = set()
        self.permissions = list(permissionMatrix.columns)
        self.userPermission = permissionMatrix.to_numpy()
        self.usernumber, self.permissionnumber = self.userPermission.shape 
        self.start = starttime

    # get permission sets of user#u                    
    def Pu(self, u):
        permissions = set()
        for i in range(len(self.userPermission[u])):
            if self.userPermission[u][i]==1:
                permissions.add(self.permissions[i])
        return permissions

    def optimizationMetric(self):
        user_role_edges = np.sum([len(role[0]) for role in self.roles])
        role_permission_edges = np.sum([len(role[2]) for role in self.roles])
        role_role_edges = np.sum([len(role[3]) for role in self.roles])
        edges_number = user_role_edges + role_permission_edges + role_role_edges
        role_number = len(self.roles)
        return edges_number + role_number
        

    def initialize(self):
        self.roles = [] 
        # role: (users, permissions, activepermissions, nextRoles)
        [ self.roles.append([set([user]), self.Pu(user), self.Pu(user), list()]) for user in range(self.usernumber)]
        self.optimization_metric = self.optimizationMetric()
        self.visited = list()
        
    def hasUnMergedRole(self):

        def hasEqualPermissions(A, B):
            return A[1] == B[1] and len(A[1])>0
        
        def hasSubsetPermissions(A, B):
            return set(B[1]).issubset(A[1])

        def hasOverlapPermission(A, B):
            return set(A[1]).intersection(B[1]) 

        for A in self.roles:
            for B in self.roles:
                if A==B:
                    continue
                if hasEqualPermissions(A, B) and (A, B) not in self.visited:
                    return 0, A, B 
                elif hasSubsetPermissions(A, B)  and (A, B) not in self.visited:
                    return 1, A, B
                elif hasOverlapPermission(A, B)  and (A, B) not in self.visited:
                    return 2, A, B 
        return -1, None, None 

    def mining(self):
        while True:
            flag, A, B = self.hasUnMergedRole()
            if flag==-1:
                break 
            if flag == 0:
                self.roles.remove(B)
                A0 = copy.deepcopy(A[0])
                A[0] = set(A[0]).union(B[0])
                new_optimizationmetric = self.optimizationMetric()
                if new_optimizationmetric > self.optimization_metric:
                    A[0] = A0 
                    self.roles.append(B)
                else:
                    self.optimization_metric = new_optimizationmetric
                self.visited.append((A, B))

            elif flag == 1:
                A2 = copy.deepcopy(A[2])
                A[2] = set(A[2]).difference(B[2])
                A[3].append(B)
                new_optimizationmetric = self.optimizationMetric()
                if new_optimizationmetric > self.optimization_metric:
                    A[2] = A2
                    A[3].remove(B)
                else:
                    self.optimization_metric = new_optimizationmetric
                self.visited.append((A, B))

            elif flag == 2:
                commonrole = (set(), A[1].intersection(B[1]), A[1].intersection(B[1]), set())
                A2 =  copy.deepcopy(A[2])
                B2 =  copy.deepcopy(B[2])
                A[2] = set(A[2]).difference(B[2])
                B[2] = set(B[2]).difference(A[2])
                self.roles.append(commonrole)
                A[3].append(commonrole)
                B[3].append(commonrole)
                new_optimizationmetric = self.optimizationMetric()
                if new_optimizationmetric > self.optimization_metric:
                    A[2] = A2
                    B[2] = B2 
                    A[3].remove(commonrole)
                    B[3].remove(commonrole)
                    self.roles.remove(commonrole)
                else:
                    self.optimization_metric = new_optimizationmetric
                self.visited.append((A, B))
    
    @timeout_decorator.timeout(20*60)                   
    def process(self):
        self.initialize()
        self.mining() 
        roles = [ set(role[1]) for role in self.roles]
        return roles 

def saveRoles2File(algo, address, timecost, roleNumber, mined_roles, labledroles, groundtruth_roles, number_ratio, roleset_roleset_sim1, roleset_roleset_sim2):
        if not os.path.exists("./roles"):
                os.mkdir("./roles")
        
        xlsx = os.path.join("./roles", "result-OpenZeppelin-HP-ORCA-HM-GO-FullBenchmark.xlsx")
        if os.path.exists(xlsx):
            wb = openpyxl.load_workbook(xlsx)
            ws = wb.active
            n = len(list(ws.rows)) + 1
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "RoleMiningResult"
            ws["A1"] = "Algorithm"
            ws["B1"] = "Address"
            ws["C1"] = "Time"
            ws["D1"] = "RoleNumber"
            ws["E1"] = "Roles"  
            ws["F1"] = "LabeledRoles"  
            ws["G1"] = "DeployedRoles"  
            ws["H1"] = "roleset_roleset_sim1"  
            ws["I1"] = "roleset_roleset_number_ratio"  
            ws["J1"] = "roleset_roleset_sim2(without permissionless)"  
            n = 2
        ws[f"A{n}"] = algo
        ws[f"B{n}"] = address
        ws[f"C{n}"] = timecost
        ws[f"D{n}"] = roleNumber
        ws[f"E{n}"] = str(mined_roles) 
        ws[f"F{n}"] = str(labledroles)
        ws[f"G{n}"] = str(groundtruth_roles)
        ws[f"H{n}"] = roleset_roleset_sim1 
        ws[f"I{n}"] = number_ratio 
        ws[f"J{n}"] = roleset_roleset_sim2 
        wb.save(xlsx)

def jaccard_func(set1, set2):
    try:
        return len(set(set1).intersection(set2))/len(set(set1).union(set2))
    except:
        return -1

def roleset_roleset_jaccard_func(roleset1, roleset2):
    
    def role_roleset_jaccard_func(role, roleset):
        max_jaccard_sim = 0
        for _role in roleset:
            jaccard_sim = jaccard_func(role, _role)
            if jaccard_sim > max_jaccard_sim:
                max_jaccard_sim = jaccard_sim
        return  max_jaccard_sim
    
    n1 = len(roleset1)
    n2 = len(roleset2)
    if n1>n2:
        roleset_tmp = roleset1
        roleset1 = roleset2
        roleset2 = roleset_tmp
    
    sum_sim = 0
    for role in roleset1:
        sum_sim = sum_sim + role_roleset_jaccard_func(role, roleset2)
    
    if len(roleset1) == 0:
        return -1
    return sum_sim/len(roleset1)

def label_func(role_permission_functions, other_permission_functions, ABIfunctions):
    rp_funcs = role_permission_functions
    op_funcs = other_permission_functions
    abi_funcs = ABIfunctions
    
    def func(observed_methods, mined_roles):
        # label mined roles with ground truth, namely rp_funcs+op_funcs
        # using jaccard similarity
        # get all relevant roles
        deployed_role_permissionset = set()
        deployed_role_labelset = set()

        relevant_roles = set(map(lambda rp: rp[1], rp_funcs)).union(set(map(lambda op: op[1], op_funcs)))
        relevant_role_permissions = dict()
        permissioned_functions = set()
        for role in relevant_roles:
            relevant_role_permissions[role] = {}
            relevant_role_permissions[role]["orig"] = set(filter(lambda rp: rp[1]==role, rp_funcs)).union(set(filter(lambda op: op[1]==role, op_funcs)))
            relevant_role_permissions[role]["simple"] =set(map(lambda p: p[0], relevant_role_permissions[role]["orig"]))
            permissioned_functions.update(relevant_role_permissions[role]["simple"])
        
        permissionless_functions = set(ABIfunctions).difference(permissioned_functions)
            
        results = list()
        for role in mined_roles:
            candidate_relevant_roles = list(filter(lambda candidate: \
                len(relevant_role_permissions[candidate]["simple"].intersection(role))>0, \
                relevant_roles))
            if len(candidate_relevant_roles)==0:
                # normal users which call only permissionless functions
                results.append(((NORMAL_USER), role, jaccard_func(role, permissionless_functions)))
                continue 
            jaccard_max_sim = 0
            label = None 
            for i in range(1, len(candidate_relevant_roles)+1):
                for role_label_pair in combinations(candidate_relevant_roles, i):
                    # calculate the jaccard similarity
                    role_label_pair_functions = set(itertools.chain.from_iterable(\
                        [ relevant_role_permissions[r]["simple"] for r in role_label_pair]))
                    # print(set(role).intersection(permissioned_functions), role_label_pair_functions)
                    if set(role).intersection(permissioned_functions).issubset(role_label_pair_functions):
                        jaccard_sim = jaccard_func(set(role).intersection(permissioned_functions), role_label_pair_functions)
                    else:
                        jaccard_sim = 0
                    if jaccard_sim > jaccard_max_sim:
                        jaccard_max_sim = jaccard_sim
                        label = role_label_pair 
                        deployed_role_labelset.update(role_label_pair)                       
            results.append((label, role, jaccard_max_sim))
      
        deployed_role_permissionset =  [ ((role_label), relevant_role_permissions[role_label]["simple"].intersection(observed_methods)) for role_label in deployed_role_labelset]
        all_role_permissionset = [ ((role_label), relevant_role_permissions[role_label]["simple"]) for role_label in relevant_roles]
        return deployed_role_labelset, deployed_role_permissionset, all_role_permissionset, results 
        
    return func

# def compare(results, groundtruth_rolepermissions):
   
#     roleset_roleset_permission_jaccard = roleset_roleset_jaccard_func([ role[1] for role in groundtruth_rolepermissions]\
#         , [ role[1] for role in results])
    
#     return roleset_roleset_permission_jaccard, len(results)/len(groundtruth_rolepermissions) if len(groundtruth_rolepermissions)>0 else -1

def jaccard_func(set1, set2):
    try:
        return len(set(set1).intersection(set2))/len(set(set1).union(set2))
    except:
        return 0

def roleset_roleset_jaccard_func_old_donot_purse_one2one(roleset1, roleset2, t=1):
    def role_roleset_jaccard_func(role, roleset):
        max_jaccard_sim = 0
        matched_role = None
        for _role in roleset:
            jaccard_sim = jaccard_func(role, _role)
            if jaccard_sim > max_jaccard_sim:
                max_jaccard_sim = jaccard_sim
                matched_role = _role
        return  max_jaccard_sim, matched_role
    
    roleset1 = copy.copy(roleset1)
    roleset2 = copy.copy(roleset2)
    n1 = len(roleset1)
    n2 = len(roleset2)
    if n1 == 0:
        return 1 

    if n1>n2:
        roleset_tmp = roleset1
        roleset1 = roleset2
        roleset2 = roleset_tmp
    
    sum_sim = 0
    matched_roles = list()
    counter = 0
    for role in roleset1:
        maxjaccardsim, matched_role = role_roleset_jaccard_func(role, roleset2)
        if matched_role is not None:
            matched_roles.append(matched_role)
        sum_sim = sum_sim + maxjaccardsim
        counter += 1
 
    for matched_role in matched_roles:
        if roleset2.count(matched_role)>0:
            roleset2.remove(matched_role)

    for restrole in roleset2:
        maxjaccardsim, _ = role_roleset_jaccard_func(restrole, roleset1)
        if maxjaccardsim > t:
            sum_sim = sum_sim + maxjaccardsim
            counter += 1

    if counter == 0:
        return -1
    return sum_sim/counter

def compareRoleSets(mined_roles, deployed_roles, t=1.0):
    print(mined_roles)
    print(deployed_roles)
    permissionless_functions = set(itertools.chain.from_iterable(mined_roles)).difference(\
        set(itertools.chain.from_iterable(deployed_roles)))
    deployed_roles.append(permissionless_functions)
    rolesroles_sim1 = roleset_roleset_jaccard_func_old_donot_purse_one2one(mined_roles, deployed_roles, t=t)
    deployed_roles.remove(permissionless_functions)
    mined_roles_remove_permissionless = [ role.difference(permissionless_functions) \
        for role in mined_roles if len(role.difference(permissionless_functions))>0]
    deployed_roles_remove_permissionless = [ role.difference(permissionless_functions) \
        for role in deployed_roles if len(role.difference(permissionless_functions))>0]
    rolesroles_sim2 = roleset_roleset_jaccard_func_old_donot_purse_one2one(mined_roles_remove_permissionless, \
        deployed_roles_remove_permissionless, t=t)
    print(rolesroles_sim1, rolesroles_sim2)
    print("*************\n")
    return  rolesroles_sim1, rolesroles_sim2



ALGO_HP = "HP"
ALGO_ORCA = "ORCA"
ALGO_GO = "GO"
ALGO_HM = "HM" 
def main():
    assert len(sys.argv[1:])==1, "shoud specify role mining algorithm (HP, GO, ORCA, HM)"
    algo = sys.argv[1]
    if algo == ALGO_HP:
        Algo = HPr
    elif algo == ALGO_GO:
        Algo = GO
    elif algo == ALGO_HM:
        Algo = HMLattice
    elif algo == ALGO_ORCA:
        Algo = ORCA
    else:
        print("unkonwn algorithm, shoud specify role mining algorithm (HP, GO, ORCA, HM)")
        exit(0)
    workdir = "OpenZeppelin1000calls10methods"
    xlsx = os.path.join("./", "OpenZeppelin1000calls10methods-label.xlsx")
    wb = openpyxl.load_workbook(xlsx, read_only=False)    
    ws = wb.active 
    contracts_limited_number = 200
    rows = list(ws.rows)[1:]
    counts = list()
    for row in rows:
        try:
            if contracts_limited_number == 0:
                break 
            start = time.time()
            address = row[0].value
            if address is None:
                    break 
            contractName = os.path.splitext(os.path.basename(row[0].hyperlink.target))[0]

            print(f"{contractName}: ", address)
            abi_file = os.path.join(os.path.dirname(row[0].hyperlink.target),contractName+".abi")
            ABIfunctions = getABIfunctions(abi_file)

            with open(os.path.join(workdir, address, "user_all.json"), "r") as f:
                    usercall_statistics = json.load(f)
                    userCount = len(usercall_statistics["data"]["ethereum"]["smartContractCalls"])
                    print("userCount:", userCount)     
                    counts.append(userCount)
                    if userCount > 100:
                        if row[3].value!="" and row[3].value!="set()":
                            role_permission_functions = eval(row[3].value)
                        else:
                            role_permission_functions = set()
                        if row[4].value!="" and row[4].value!="set()":
                            other_permission_functions = eval(row[4].value)
                        else:
                            other_permission_functions = set()
                        # print(role_permission_functions, other_permission_functions)
                        eval_func = label_func(role_permission_functions, other_permission_functions, ABIfunctions)
                        
                        observed_methods, verybasic_userGroups ,permissionMatrix = lightweightrolemining(address=address, workdir=workdir)
                        end = time.time()
                        verybasic_roles = [usergroup[1] for usergroup in verybasic_userGroups]
                        deployed_role_labelset, deployed_role_permissionset, all_role_permissionset, _ = eval_func(observed_methods, verybasic_roles)
                        deployed_roles = [ set(role[1]) for role in deployed_role_permissionset]
                       
                        starttime=time.time()
                        mined_roles = Algo(permissionMatrix=permissionMatrix, starttime=starttime).process()
                        timecost = time.time()-starttime
                        _, _, _, labeled_mined_roles = eval_func(observed_methods, mined_roles)

                        mined_roles = [ set(role[1]) if isinstance(role, tuple) else role for role in mined_roles]
                        
                        number_ratio = len(mined_roles)/(len(deployed_roles)+1)

                        rolesim1, rolesim2 = compareRoleSets(mined_roles=mined_roles, deployed_roles=deployed_roles)
                      
                        saveRoles2File(algo, address, timecost,  len(mined_roles), mined_roles, labeled_mined_roles, deployed_role_permissionset, rolesim1,
                        number_ratio, rolesim2)

                        contracts_limited_number -= 1
        except:
            traceback.print_exc()
            continue  

if __name__ == "__main__":
    main()
